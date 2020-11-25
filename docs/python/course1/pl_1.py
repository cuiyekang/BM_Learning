
# filepath="./docs/python/course1/t_data/szzs.txt"

# date =[]
# pct =[]

# with open(filepath) as filehandle:
#     for line in filehandle:
#         a,b =line.split()
#         date.append(a)
#         pct.append(b)


# filepath="./docs/python/course1/t_data/pl_write.txt"

# with open(filepath,'w') as filehandle:
#     filehandle.write("test")
#     filehandle.close()

# with open(filepath,'a') as filehandle:
#     filehandle.write("test1 \n")
#     filehandle.write("test2 \n")
#     filehandle.close()

import openpyxl

# wb = openpyxl.Workbook()
# ws = wb.active
# ws1 = wb.create_sheet("Mysheel")
# ws.Title = "New Title"
# ws['A4']=4
# cell_range = ws['A1':'C2']
# for cell in cell_range:
#     for i in cell:
#         i.value = 2

# wb.save("./docs/python/course1/t_data/pl_1_write.xls")
# wb.close

wb = openpyxl.load_workbook("./docs/python/course1/t_data/pl_1_1.xlsx")
print(wb.sheetnames)
ws = wb["profits"]

dict = {}

for row in ws.iter_rows(min_row=2,max_col=5,max_row=len(ws['D'])):
    dict[row[3].value] = dict.get(row[3].value,0) + row[4].value


use_range = ws['H1':'I3602']

for i in use_range:
    if dict.get(i[0].value,0) != 0:
        i[1].value = dict.get(i[0].value)

wb.save("./docs/python/t_data/course1/pl_1_1_finish.xlsx")
wb.close
